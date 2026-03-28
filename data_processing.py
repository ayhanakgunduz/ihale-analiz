import pandas as pd
import numpy as np
import re

def clean_monetary_value(val):
    """
    Takes an arbitrary Excel cell value (text or number) and aggressively cleans it
    into a float. Handles Turkish currency formats (1.234,56 TL) and mixed types.
    """
    if pd.isna(val):
        return np.nan
    
    if isinstance(val, (int, float)):
        return float(val)

    val_str = str(val).strip().upper()
    
    # Kötü veriler veya "yok", "teklif verilmedi" gibi durumlar
    # Farklı tire (dash) karakterlerini ve yaygın boş değer ifadelerini ekleyelim
    if val_str in ['-', '–', '—', '', 'YOK', 'N/A', 'NULL', 'NA', '.', '...']:
        return np.nan

    # Metni tamamen soyutla: Sadece rakam, virgül, nokta veya eksi işareti kalsın
    # Örn: "1.234,56 TL" -> "1.234,56"
    val_str = re.sub(r'[^\d.,-]', '', val_str)
    
    if not val_str:
        return np.nan

    # Türkiye/Avrupa formatı (nokta binlik, virgül ondalık) vs Amerikan formatı
    # Önce virgül ve nokta sayısına bakalım
    count_comma = val_str.count(',')
    count_dot = val_str.count('.')
    
    # 1.234.567,89 durumu
    if count_comma == 1 and count_dot >= 0:
        # Eğer virgül sondan 2. veya 3. karakterse ondalık ayracı olma ihtimali çok yüksektir
        if len(val_str) - val_str.find(',') <= 4:
            val_str = val_str.replace('.', '')  # Binlik ayraçlarını sil
            val_str = val_str.replace(',', '.') # Ondalık ayracını noktaya çevir
        else:
            # Virgül binlik ayırıcısı olabilir mi? (Amerikan formatı 1,234.50)
            if count_dot <= 1:
                val_str = val_str.replace(',', '')
    # 1,234,567.89 (Amerikan standart biçimi)
    elif count_dot == 1 and count_comma >= 0:
        val_str = val_str.replace(',', '')
    else:
        # Birden fazla veya hiç yoksa sadece nokta veya sadece virgül vardır
        # Eğer sadece tek bir ayrac varsa ve o da virgülese, onu noktaya çevirelim
        if count_comma == 1 and count_dot == 0:
            val_str = val_str.replace(',', '.')
        # Birden çok ayraç varsa (1.234.567), ondalık olmadığını varsayıp hepsini sil
        elif count_dot > 1 and count_comma == 0:
            val_str = val_str.replace('.', '')
        elif count_comma > 1 and count_dot == 0:
            val_str = val_str.replace(',', '')

    try:
        return float(val_str)
    except ValueError:
        return np.nan

def load_and_clean_data(file, raw_sheet_name, params_sheet_name):
    """
    Excel dosyasını okur, belirtilen sayfaları çeker ve dinamik temizlik yapar.
    Sözlük döndürür: {'raw': df_raw, 'params': df_params, 'organized': df_org}
    """
    # Akıllı Başlık Tespiti için sayfayı önce ham halde okuyalım
    try:
        # Header=None ile okuyarak satırları manuel tarayacağız
        df_full = pd.read_excel(file, sheet_name=raw_sheet_name, header=None)
    except Exception as e:
        raise ValueError(f"Raw data '{raw_sheet_name}' okunamadı: {e}")

    # Başlığı bulmak için ilk 20 satırı tara
    header_row_idx = 0
    keywords = ['no', 'iş kalemi', 'is kalemi', 'açıklama', 'aciklama', 'description', 'tem no', 'kalem no', 'item no']
    
    # Satır satır tarayalım
    for i in range(min(len(df_full), 20)):
        row_values = [str(val).lower() for val in df_full.iloc[i].values if pd.notna(val)]
        # Eğer bu satırda anahtar kelimelerden en az 2 tanesi varsa bu başlık satırıdır
        if sum(1 for kw in keywords if any(kw in val for val in row_values)) >= 2:
            header_row_idx = i
            break
            
    # Tespit edilen satırı başlık yapalım ve altındaki verileri alalım
    df_raw = df_full.iloc[header_row_idx:].reset_index(drop=True)
    df_raw.columns = df_raw.iloc[0] # İlk satırı kolon yap
    df_raw = df_raw.iloc[1:].reset_index(drop=True) # Başlık satırını veriden çıkar

    # Tamamen NaN olan satır/sütunları düşelim (Bu sol taraftaki boş A sütununu da temizler)
    df_raw = df_raw.dropna(how='all', axis=0).dropna(how='all', axis=1)
    
    # Diğer sayfaları da okuyalım (Varsa)
    try:
        if params_sheet_name:
            df_params = pd.read_excel(file, sheet_name=params_sheet_name)
        else:
            df_params = pd.DataFrame()
    except:
        df_params = pd.DataFrame()

    # 1. Kolon Tiplerini Tespit Et (Metadata vs Firma)
    metadata_keywords = keywords + [
        'birim', 'unit of measure', 'uom', 'miktar', 'quantity', 'qty',
        'poz', 'malzeme', 'commodity', 'supplier', 'tedarikçi', 'tedarikci',
        'firma adı', 'firma adi', 'company', 'not', 'note', 'kategori', 'category'
    ]
    
    firm_cols = []
    meta_cols = []
    
    for col in df_raw.columns:
        col_lower = str(col).lower()
        if any(kw in col_lower for kw in metadata_keywords):
            meta_cols.append(col)
        else:
            firm_cols.append(col)
    
    # 2. Döviz Birimini Tespit Et (Önce Excel hücre formatından, sonra metin içeriğinden)
    currency_symbol = "₺" # Varsayılan
    currency_found = False
    
    # 2a. openpyxl ile Excel hücre number_format'ını kontrol et
    currency_format_map = {
        'RUB': '₽', 'РУБ': '₽', '₽': '₽',
        'USD': '$', '$': '$',
        'EUR': '€', '€': '€',
        'GBP': '£', '£': '£',
        'AED': 'AED', 'SAR': 'SAR', 'TL': '₺', 'TRY': '₺',
    }
    try:
        import openpyxl
        if hasattr(file, 'seek'): file.seek(0)
        wb = openpyxl.load_workbook(file, data_only=True, read_only=True)
        ws = wb[raw_sheet_name] if raw_sheet_name in wb.sheetnames else wb.active
        # İlk 10 satırdaki sayısal hücrelerin formatlarını tara
        for row in ws.iter_rows(min_row=2, max_row=12, min_col=1, max_col=ws.max_column):
            for cell in row:
                fmt = str(cell.number_format).upper() if cell.number_format else ''
                if fmt and fmt != 'GENERAL':
                    for key, sym in currency_format_map.items():
                        if key in fmt:
                            currency_symbol = sym
                            currency_found = True
                            break
                if currency_found: break
            if currency_found: break
        wb.close()
        if hasattr(file, 'seek'): file.seek(0)
    except Exception:
        pass  # openpyxl başarısız olursa metin taramasına geç
    
    # 2b. Fallback: Hücre içeriğinde para birimi sembolü ara
    if not currency_found:
        sample_cells = df_raw.tail(20).astype(str).values.flatten()
        for symbol in ['$', '€', '£', 'RUB', '₽', 'AED', 'SAR', 'РУБ']:
            if any(symbol in str(c).upper() for c in sample_cells):
                currency_symbol = '₽' if symbol in ['RUB', 'РУБ', '₽'] else symbol
                currency_found = True
                break

    # 3. "Grand Total" Satırını Daha Hassas Tespit Et
    summary_keywords = ['grand total', 'genel toplam', 'toplam bedel', 'ara toplam', 'grandtotal']
    
    # Sadece ilk 5 kolonda aramak "Description" içindeki rastgele "toplam"ları eler
    check_cols = df_raw.columns[:min(5, len(df_raw.columns))]
    mask = df_raw[check_cols].astype(str).apply(
        lambda row: row.str.lower().str.contains('|'.join(summary_keywords), na=False).any(), 
        axis=1
    )
    
    excel_totals = {}
    if mask.any():
        total_rows = df_raw[mask]
        # Long-format: Her Grand Total satırında Supplier ve Total Price/Unit Price olabilir
        sup_col_name = next((c for c in df_raw.columns if any(kw in str(c).lower() for kw in ['supplier', 'tedarikçi', 'tedarikci', 'firma adı', 'firma adi'])), None)
        total_price_col = next((c for c in df_raw.columns if 'total price' in str(c).lower() or 'toplam fiyat' in str(c).lower()), None)
        
        if sup_col_name and total_price_col:
            # Her firma için Grand Total değerini al
            for _, trow in total_rows.iterrows():
                supplier = str(trow.get(sup_col_name, '')).strip()
                total_val = clean_monetary_value(trow.get(total_price_col))
                if supplier and pd.notna(total_val) and total_val > 0:
                    excel_totals[supplier] = total_val
        else:
            # Eski yöntem (wide-format fallback)
            total_row = total_rows.iloc[-1]
            for col in firm_cols:
                if col in total_row:
                    val = clean_monetary_value(total_row[col])
                    if pd.notna(val):
                        excel_totals[str(col)] = val
                
    # 4. Organize Veriyi Oluştur ve Özet Satırları Temizle
    df_org = df_raw[~mask].reset_index(drop=True)
            
    # Eğer hiç meta column bulunamadıysa, ilk 2 kolonu metadata kabul edelim
    # Dinamik Pivot Kontrolü (Long-to-Wide Dönüşümü)
    # Eğer "Supplier" (veya benzeri) bir kolon varsa ve "Unit Price" kolonu da varsa pivot yap.
    # -----------------------------------------------------------------------------------
    sup_col = next((c for c in df_org.columns if any(kw in str(c).lower() for kw in ['supplier', 'tedarikçi', 'tedarikci', 'firma adı', 'firma adi'])), None)
    price_col = next((c for c in df_org.columns if any(kw in str(c).lower() for kw in ['unit price', 'birim fiyat', 'teklif', 'price'])), None)
    
    if sup_col and price_col:
        # Pivat için index kolonlarını bul (No, Description vb.)
        index_cols = [c for c in df_org.columns if any(kw in str(c).lower() for kw in ['no', 'kalem', 'birim', 'description', 'açıklama', 'miktar'])]
        # Bazı kolonları listeden çıkar (Pivot yapılacaklar)
        index_cols = [c for c in index_cols if c != sup_col and c != price_col]
        
        try:
            # Sadece sayısal kısımları temizle
            df_org[price_col] = pd.to_numeric(df_org[price_col].apply(clean_monetary_value), errors='coerce')
            
            # Pivot İşlemi: Item No ve Description sabit kalsın, Supplier'lar başlık olsun, Price'lar değer.
            df_pivot = df_org.pivot_table(
                index=index_cols, 
                columns=sup_col, 
                values=price_col, 
                aggfunc='first'
            ).reset_index()
            
            # Pivot başarılı olduysa firmaları güncelle
            df_org = df_pivot
            firm_cols = [c for c in df_org.columns if c not in index_cols]
            meta_cols = index_cols
        except:
            # Pivot başarısız olursa eski yöntemle devam et
            pass

    # Okunan firmaları sayısal formata zorla
    for col in firm_cols:
        df_org[col] = pd.to_numeric(df_org[col].apply(clean_monetary_value), errors='coerce')
    
    # Arrow-uyumluluğu için tüm meta kolonlarında tire ('-') karakterlerini temizleyelim
    for col in df_org.columns:
        if df_org[col].dtype == 'object':
            # Tirelerin Arrow tarafından double sütunlarda 'str' olarak algılanmasını önle
            df_org[col] = df_org[col].replace(['-', '–', '—'], np.nan)
            
            # Eğer kolon hala objeyse ve çoğunlukla sayısal gözüküyorsa (örn: 'Unit Price'), zorla
            try:
                sample = df_org[col].dropna().head(10).astype(str)
                if all(re.match(r'^[\d\.,\s\-]*$', s) for s in sample if s):
                    cleaned = df_org[col].apply(clean_monetary_value)
                    if cleaned.notna().sum() > 0:
                        df_org[col] = pd.to_numeric(cleaned, errors='coerce')
            except:
                pass

    # Sütunları kullanıcı isteğine göre yeniden sırala (1. Item No, 2. Description, 3. Firmalar)
    # -----------------------------------------------------------------------------------
    final_cols = []
    
    # 1. Item No / No / Tem No
    no_col = next((c for c in df_org.columns if any(kw in str(c).lower() for kw in ['tem no', 'kalem no', 'item no', ' no', 'no.'])), None)
    if not no_col: no_col = next((c for c in df_org.columns if str(c).lower() == 'no'), None)
    if no_col: final_cols.append(no_col)
    
    # 2. Description / Açıklama / Kalem Adı
    desc_col = next((c for c in df_org.columns if any(kw in str(c).lower() for kw in ['description', 'açıklama', 'aciklama', 'kalem adı', 'malzeme'])), None)
    if desc_col and desc_col not in final_cols: final_cols.append(desc_col)
    
    # 3. Firma Teklifleri (Birim Fiyatlar)
    for c in firm_cols:
        if c not in final_cols:
            final_cols.append(c)
            
    # Kalan tüm metadata kolonları (Supplier, Miktar, Birim vs) sona ekle
    for c in meta_cols:
        if c not in final_cols:
            final_cols.append(c)
            
    # Eğer kolonlar mevcutsa yeniden sırala (Eksikse skip atar)
    df_org = df_org[[c for c in final_cols if c in df_org.columns]]

    # Miktar/Quantity kolonunu temizle
    qty_col = next((c for c in df_org.columns if any(kw in str(c).lower() for kw in ['miktar', 'quantity', 'qty', 'adet'])), None)
    if qty_col:
        def extract_qty(val):
            if pd.isna(val): return 1.0
            val_str = str(val).strip()
            import re
            m = re.search(r'([\d\.\,]+)', val_str)
            if m:
                cleaned = clean_monetary_value(m.group(1))
                return cleaned if pd.notna(cleaned) else 1.0
            return 1.0
        df_org[qty_col] = df_org[qty_col].apply(extract_qty)

    # Params sheet temizliği
    if not df_params.empty:
        for col in df_params.columns:
            if df_params[col].dtype == 'object':
                df_params[col] = df_params[col].replace(['-', '–', '—'], np.nan)
                # Sayısal olabilecek sütunları otomatik tespit edip çevirelim
                df_params[col] = pd.to_numeric(df_params[col], errors='ignore')
    
    # Aykırı değerler sayfası (Z-skoru eşiği vb. içerebilir)
    try:
        # 'AYKIRI DEGERLER' veya 'OUTLIERS' sayfası var mı bak
        outlier_sheets = [s for s in pd.ExcelFile(file).sheet_names if 'aykırı' in s.lower() or 'outlier' in s.lower()]
        if outlier_sheets:
            df_outliers_params = pd.read_excel(file, sheet_name=outlier_sheets[0])
        else:
            df_outliers_params = pd.DataFrame()
    except:
        df_outliers_params = pd.DataFrame()
    
    # 5. Numerik Sıralama (1, 2, 3... 10, 11... sırasını garanti altına al)
    if no_col and no_col in df_org.columns:
        try:
            # Geçici bir kolon kullanarak numerik sıralama yapalım (orijinal veriyi bozmadan)
            df_org = df_org.sort_values(
                by=no_col, 
                key=lambda x: pd.to_numeric(x.astype(str).str.replace(r'[^\d]', '', regex=True), errors='coerce'),
                ascending=True
            ).reset_index(drop=True)
        except:
            pass # Sıralama başarısız olursa orijinal sırayı koru
            
    return {
        'raw': df_raw,
        'params': df_params,
        'outliers_params': df_outliers_params,
        'organized': df_org,
        'meta_cols': meta_cols,
        'firm_cols': firm_cols,
        'excel_totals': excel_totals,
        'currency': currency_symbol
    }
